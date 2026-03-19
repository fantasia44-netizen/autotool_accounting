"""주문/출고/피킹/패킹 관련 라우트."""
import logging
from flask import render_template, request, redirect, url_for, flash
from flask_login import login_required, current_user

from . import operator_bp, _require_operator

logger = logging.getLogger(__name__)


# ═══ 주문관리 ═══

@operator_bp.route('/orders')
@login_required
@_require_operator
def orders():
    from db_utils import get_repo
    repo = get_repo('order')
    status = request.args.get('status')
    channel = request.args.get('channel')
    client_id = request.args.get('client_id', type=int)

    orders_list = repo.list_orders(status=status, channel=channel,
                                   client_id=client_id) or []
    counts = repo.count_by_status() or {}

    client_repo = get_repo('client')
    clients = client_repo.list_clients() or []
    client_map = {c['id']: c['name'] for c in clients}

    return render_template('operator/orders.html', orders=orders_list, counts=counts,
                           filter_status=status, filter_channel=channel,
                           filter_client_id=client_id,
                           clients=clients, client_map=client_map)


@operator_bp.route('/orders/<int:order_id>')
@login_required
@_require_operator
def order_detail(order_id):
    from db_utils import get_repo
    repo = get_repo('order')
    order = repo.get_order_with_items(order_id)
    if not order:
        flash('주문을 찾을 수 없습니다.', 'warning')
        return redirect(url_for('operator.orders'))

    status_logs = repo.get_status_logs(order_id) or []
    shipments = repo.list_shipments(order_id=order_id) or []
    return render_template('operator/order_detail.html', order=order,
                           status_logs=status_logs, shipments=shipments)


@operator_bp.route('/orders/<int:order_id>/status', methods=['POST'])
@login_required
@_require_operator
def order_status_update(order_id):
    from db_utils import get_repo
    repo = get_repo('order')
    order = repo.get_order(order_id)
    old_status = order.get('status', '') if order else ''
    new_status = request.form.get('status')
    reason = request.form.get('reason', '').strip()
    # 재고 예약/해제 연동
    inv_repo = get_repo('inventory')
    if new_status == 'confirmed' and old_status in ('pending', ''):
        # 풀필먼트 모드 판별 + 단품/합포 분류
        try:
            from services.fulfillment_mode_service import determine_order_mode
            client_repo = get_repo('client')
            client = client_repo.get_client(order.get('client_id')) or {}
            order_full = repo.get_order_with_items(order_id)
            items = order_full.get('items', []) if order_full else []
            inv_repo_tmp = get_repo('inventory')
            sku_ids = [it.get('sku_id') for it in items if it.get('sku_id')]
            sku_map = {}
            for sid in sku_ids:
                sku = inv_repo_tmp.get_sku(sid)
                if sku:
                    sku_map[sid] = sku
            mode_result = determine_order_mode(client, items, sku_map)
            repo.update_order(order_id, {
                'fulfillment_mode': mode_result['mode'],
                'pack_type': mode_result['pack_type'],
            })
            if mode_result.get('downgraded'):
                flash(f'안정모드 SKU 포함 → 전체 안정모드(B)로 강등', 'info')
        except Exception as e:
            current_app.logger.warning(f'[모드판별실패] order_id={order_id}: {e}')

        from services.inventory_service import reserve_stock
        result = reserve_stock(inv_repo, repo, order_id)
        if not result.get('ok'):
            flash(f'재고 예약 실패: {result.get("error", "알 수 없는 오류")}', 'danger')
            return redirect(url_for('operator.order_detail', order_id=order_id))

    if new_status == 'cancelled' and old_status not in ('cancelled', ''):
        from services.inventory_service import release_stock
        release_stock(inv_repo, order_id)

    repo.update_order_status(order_id, new_status)
    repo.log_status_change(order_id, old_status, new_status,
                           changed_by=current_user.id, reason=reason)
    # 출고 시 과금 기록 (서비스 내부에서 DLQ 처리)
    if new_status == 'shipped' and order:
        try:
            cid = order.get('client_id')
            if cid:
                from services.client_billing_service import record_outbound_fee
                # 주문 아이템 수 / 총 중량 계산 → 조건부 과금
                order_full = repo.get_order_with_items(order_id)
                items = order_full.get('items', []) if order_full else []
                item_count = len(items)
                total_weight_g = sum(
                    (it.get('weight_g', 0) or 0) * abs(it.get('quantity', it.get('qty', 1)))
                    for it in items
                )
                record_outbound_fee(get_repo('client_billing'),
                                    get_repo('client_rate'), cid,
                                    order_id=order_id,
                                    item_count=item_count,
                                    total_weight_g=total_weight_g)
        except Exception:
            logger.exception('과금 서비스 호출 자체 실패 (출고): order_id=%s', order_id)
    flash(f'주문 상태가 "{new_status}"로 변경되었습니다.', 'success')
    return redirect(url_for('operator.order_detail', order_id=order_id))


@operator_bp.route('/orders/<int:order_id>/hold', methods=['POST'])
@login_required
@_require_operator
def order_hold(order_id):
    """주문 보류 처리."""
    from db_utils import get_repo
    repo = get_repo('order')
    reason = request.form.get('reason', '').strip() or '관리자 보류'
    order = repo.get_order(order_id)
    if order:
        repo.hold_order(order_id, reason, current_user.id)
        repo.log_status_change(order_id, order.get('status', ''), 'hold',
                               changed_by=current_user.id, reason=reason)
        # 보류 시 예약 해제
        if order.get('status') in ('confirmed', 'picking_ready', 'picking'):
            try:
                inv_repo = get_repo('inventory')
                from services.inventory_service import release_stock
                release_stock(inv_repo, order_id)
            except Exception:
                logger.exception('재고 예약 해제 실패: order_id=%s', order_id)
    flash(f'주문이 보류되었습니다: {reason}', 'warning')
    return redirect(url_for('operator.order_detail', order_id=order_id))


@operator_bp.route('/orders/<int:order_id>/release-hold', methods=['POST'])
@login_required
@_require_operator
def order_release_hold(order_id):
    from db_utils import get_repo
    repo = get_repo('order')
    order = repo.get_order(order_id)
    if order:
        repo.release_hold(order_id)
        repo.log_status_change(order_id, 'hold', order.get('status', ''),
                               changed_by=current_user.id, reason='보류 해제')
    flash('보류가 해제되었습니다.', 'success')
    return redirect(url_for('operator.order_detail', order_id=order_id))


# ═══ 피킹관리 ═══

@operator_bp.route('/picking')
@login_required
@_require_operator
def picking():
    """피킹리스트 관리."""
    from db_utils import get_repo
    picking_repo = get_repo('picking')
    order_repo = get_repo('order')
    client_repo = get_repo('client')
    wh_repo = get_repo('warehouse')

    status = request.args.get('status')
    lists = picking_repo.list_picking_lists(status=status) or []
    counts = picking_repo.count_by_status() or {}

    confirmed_orders = order_repo.list_orders(status='confirmed', limit=100) or []
    clients = client_repo.list_clients() or []
    warehouses = wh_repo.list_warehouses() or []

    return render_template('operator/picking.html', lists=lists, counts=counts,
                           confirmed_orders=confirmed_orders,
                           clients=clients, warehouses=warehouses,
                           filter_status=status)


@operator_bp.route('/picking/generate', methods=['POST'])
@login_required
@_require_operator
def picking_generate():
    """피킹리스트 생성."""
    from db_utils import get_repo
    order_ids_raw = request.form.getlist('order_ids')
    order_ids = [int(x) for x in order_ids_raw if x.isdigit()]
    list_type = request.form.get('list_type', 'by_order')
    warehouse_id = request.form.get('warehouse_id', type=int)
    client_id = request.form.get('client_id', type=int)

    if not order_ids:
        flash('주문을 선택해주세요.', 'warning')
        return redirect(url_for('operator.picking'))

    # 속도모드 주문 여부 확인 → 자동 분기
    order_repo = get_repo('order')
    use_speed = False
    for oid in order_ids:
        o = order_repo.get_order(oid)
        if o and o.get('fulfillment_mode') == 'speed':
            use_speed = True
            break

    try:
        if use_speed:
            from services.picking_service import generate_speed_picking
            pl = generate_speed_picking(
                picking_repo=get_repo('picking'),
                order_repo=order_repo,
                inv_repo=get_repo('inventory'),
                order_ids=order_ids,
                warehouse_id=warehouse_id,
                client_id=client_id,
                created_by=current_user.id,
            )
            summary = pl.get('speed_summary', {})
            flash(f'속도모드 피킹리스트 {pl.get("list_no", "")} 생성 — '
                  f'SKU {summary.get("total_skus", 0)}종 / '
                  f'단품 {summary.get("single_orders", 0)}건 / '
                  f'합포 {summary.get("multi_orders", 0)}건', 'success')
        else:
            from services.picking_service import generate_picking_list
            pl = generate_picking_list(
                picking_repo=get_repo('picking'),
                order_repo=order_repo,
                inv_repo=get_repo('inventory'),
                wh_repo=get_repo('warehouse'),
                order_ids=order_ids,
                warehouse_id=warehouse_id,
                client_id=client_id,
                list_type=list_type,
                created_by=current_user.id,
            )
            flash(f'피킹리스트 {pl.get("list_no", "")} 생성 완료 ({len(pl.get("items", []))}건)', 'success')
    except Exception as e:
        flash(f'피킹리스트 생성 오류: {e}', 'danger')

    return redirect(url_for('operator.picking'))


@operator_bp.route('/picking/<int:list_id>')
@login_required
@_require_operator
def picking_detail(list_id):
    """피킹리스트 상세."""
    from db_utils import get_repo
    repo = get_repo('picking')
    pl = repo.get_picking_list_with_items(list_id)
    if not pl:
        flash('피킹리스트를 찾을 수 없습니다.', 'warning')
        return redirect(url_for('operator.picking'))

    inv_repo = get_repo('inventory')
    sku_map = {}
    for item in pl.get('items', []):
        sid = item.get('sku_id')
        if sid and sid not in sku_map:
            sku = inv_repo.get_sku(sid)
            sku_map[sid] = sku.get('name', f'SKU#{sid}') if sku else f'SKU#{sid}'

    return render_template('operator/picking_detail.html', pl=pl, sku_map=sku_map)


@operator_bp.route('/picking/<int:list_id>/complete', methods=['POST'])
@login_required
@_require_operator
def picking_complete(list_id):
    """피킹리스트 완료."""
    from db_utils import get_repo
    repo = get_repo('picking')
    repo.complete_picking_list(list_id)
    flash('피킹리스트가 완료 처리되었습니다.', 'success')
    return redirect(url_for('operator.picking_detail', list_id=list_id))


# ═══ 택배출고 ═══

@operator_bp.route('/shipments')
@login_required
@_require_operator
def shipments():
    """택배출고 목록."""
    from db_utils import get_repo
    repo = get_repo('order')
    status = request.args.get('status')
    items = repo.list_shipments(status=status, shipment_type='normal')
    counts = repo.count_shipments_by_status()
    client_repo = get_repo('client')
    clients = client_repo.list_clients() or []
    client_map = {c['id']: c['name'] for c in clients}

    # order_id → client_id 매핑 (1회 bulk 조회)
    order_ids = list({s['order_id'] for s in items if s.get('order_id')})
    if order_ids:
        order_rows = repo._query(repo.ORDER_TABLE,
                                  columns='id,client_id,order_no',
                                  filters=[('id', 'in', order_ids)],
                                  limit=len(order_ids))
        order_client = {o['id']: o.get('client_id') for o in order_rows}
        for s in items:
            if s.get('order_id') and not s.get('client_id'):
                s['client_id'] = order_client.get(s['order_id'])

    return render_template('operator/shipments.html', shipments=items, counts=counts,
                           filter_status=status, client_map=client_map)


# ═══ 택배 엑셀 다운로드/업로드 ═══

@operator_bp.route('/shipments/excel-download')
@login_required
@_require_operator
def shipment_excel_download():
    """출고 대상 주문 엑셀 다운로드 (택배사 전달용).

    packed 상태 주문을 엑셀로 내보내기.
    컬럼: 주문번호, 수취인, 연락처, 주소, 우편번호, 상품명, 수량, 메모, 택배사, 송장번호(빈칸)
    """
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from flask import send_file
    from db_utils import get_repo

    repo = get_repo('order')
    inv_repo = get_repo('inventory')
    client_repo = get_repo('client')

    # packed 또는 confirmed 상태 주문 조회
    status = request.args.get('status', 'packed')
    orders = repo.list_orders(status=status, limit=500) or []

    if not orders:
        flash(f'"{status}" 상태 주문이 없습니다.', 'warning')
        return redirect(url_for('operator.shipments'))

    # SKU 캐시
    all_skus = inv_repo.list_skus() or []
    sku_map = {s['id']: s for s in all_skus}
    clients = client_repo.list_clients() or []
    client_map = {c['id']: c['name'] for c in clients}

    wb = Workbook()
    ws = wb.active
    ws.title = '출고대상'

    # 헤더
    headers = ['주문번호', '고객사', '채널', '수취인', '연락처', '주소', '우편번호',
               '상품명', '수량', '메모', '택배사', '송장번호']
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # 데이터
    row_num = 2
    for order in orders:
        items = repo.get_order_items(order['id']) or []
        product_names = []
        total_qty = 0
        for item in items:
            sku = sku_map.get(item.get('sku_id'), {})
            name = sku.get('name', f"SKU#{item.get('sku_id', '?')}")
            qty = item.get('quantity', item.get('qty', 1))
            product_names.append(f"{name}({qty})")
            total_qty += qty

        ws.cell(row=row_num, column=1, value=order.get('order_no', ''))
        ws.cell(row=row_num, column=2, value=client_map.get(order.get('client_id'), ''))
        ws.cell(row=row_num, column=3, value=order.get('channel', ''))
        ws.cell(row=row_num, column=4, value=order.get('recipient_name', ''))
        ws.cell(row=row_num, column=5, value=order.get('recipient_phone', ''))
        ws.cell(row=row_num, column=6, value=order.get('recipient_address', ''))
        ws.cell(row=row_num, column=7, value=order.get('zipcode', ''))
        ws.cell(row=row_num, column=8, value=', '.join(product_names))
        ws.cell(row=row_num, column=9, value=total_qty)
        ws.cell(row=row_num, column=10, value=order.get('memo', ''))
        ws.cell(row=row_num, column=11, value='')  # 택배사 (빈칸)
        ws.cell(row=row_num, column=12, value='')  # 송장번호 (빈칸)
        row_num += 1

    # 열 너비 자동 조절
    col_widths = [15, 12, 10, 10, 15, 40, 10, 30, 8, 20, 12, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # 숨겨진 order_id 컬럼 (업로드 시 매칭용)
    ws.cell(row=1, column=13, value='order_id')
    ws.column_dimensions['M'].hidden = True
    for i, order in enumerate(orders, 2):
        ws.cell(row=i, column=13, value=order['id'])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    from services.tz_utils import today_kst
    filename = f'출고대상_{today_kst()}_{status}.xlsx'
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


@operator_bp.route('/shipments/invoice-upload', methods=['POST'])
@login_required
@_require_operator
def shipment_invoice_upload():
    """송장 엑셀 업로드 — 주문에 택배사/송장번호 매칭 + shipped 전환.

    엑셀 컬럼: 주문번호, ..., 택배사, 송장번호, order_id(숨김)
    송장번호가 있는 행만 처리.
    """
    from db_utils import get_repo
    from openpyxl import load_workbook
    from io import BytesIO

    file = request.files.get('invoice_file')
    if not file or not file.filename.endswith(('.xlsx', '.xls')):
        flash('엑셀 파일(.xlsx)을 선택해주세요.', 'warning')
        return redirect(url_for('operator.shipments'))

    try:
        wb = load_workbook(BytesIO(file.read()), read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        flash(f'엑셀 파일 읽기 오류: {e}', 'danger')
        return redirect(url_for('operator.shipments'))

    repo = get_repo('order')
    inv_repo = get_repo('inventory')

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    success = 0
    skipped = 0
    errors = []

    for row_idx, row in enumerate(rows, 2):
        if len(row) < 12:
            continue

        order_no = str(row[0] or '').strip()
        carrier = str(row[10] or '').strip()  # 택배사 (col K)
        invoice_no = str(row[11] or '').strip()  # 송장번호 (col L)
        order_id = row[12] if len(row) > 12 else None  # 숨겨진 order_id (col M)

        if not invoice_no:
            skipped += 1
            continue

        # order_id로 먼저 매칭, 없으면 order_no로
        order = None
        if order_id:
            try:
                order = repo.get_order(int(order_id))
            except (ValueError, TypeError):
                pass
        if not order and order_no:
            found = repo._query(repo.ORDER_TABLE,
                                filters=[('order_no', 'eq', order_no)],
                                limit=1)
            order = found[0] if found else None

        if not order:
            errors.append(f'행 {row_idx}: 주문 "{order_no}" 없음')
            continue

        if order.get('status') == 'shipped':
            skipped += 1
            continue

        # shipment 생성 + 주문 shipped 전환
        try:
            repo.create_shipment({
                'order_id': order['id'],
                'shipment_type': 'normal',
                'invoice_no': invoice_no,
                'status': 'shipped',
                'client_id': order.get('client_id'),
            })

            old_status = order.get('status', '')
            repo.update_order_status(order['id'], 'shipped')
            repo.log_status_change(order['id'], old_status, 'shipped',
                                   changed_by=current_user.id,
                                   reason=f'송장업로드: {carrier} {invoice_no}')

            # 출고비 과금
            try:
                cid = order.get('client_id')
                if cid:
                    from services.client_billing_service import record_outbound_fee
                    order_full = repo.get_order_with_items(order['id'])
                    items = order_full.get('items', []) if order_full else []
                    record_outbound_fee(get_repo('client_billing'),
                                        get_repo('client_rate'), cid,
                                        order_id=order['id'],
                                        item_count=len(items))
            except Exception:
                logger.exception('송장업로드 과금 실패: order_id=%s', order['id'])

            success += 1
        except Exception as e:
            errors.append(f'행 {row_idx}: {order_no} 처리 실패 ({e})')

    msg = f'송장 업로드 완료: {success}건 처리'
    if skipped:
        msg += f', {skipped}건 스킵'
    if errors:
        msg += f', {len(errors)}건 오류'
        for err in errors[:5]:  # 최대 5개만 표시
            flash(err, 'warning')
    flash(msg, 'success' if not errors else 'warning')
    return redirect(url_for('operator.shipments'))


# ═══ 반품관리 (고객 회송) ═══

@operator_bp.route('/returns')
@login_required
@_require_operator
def returns():
    """반품관리 목록."""
    from db_utils import get_repo
    repo = get_repo('order')
    status = request.args.get('status')
    items = repo.list_shipments(status=status, shipment_type='return')
    client_repo = get_repo('client')
    clients = client_repo.list_clients() or []
    client_map = {c['id']: c['name'] for c in clients}
    wh_repo = get_repo('warehouse')
    warehouses = wh_repo.list_warehouses() or []
    locations = wh_repo.list_all_locations_with_path()
    inv_repo = get_repo('inventory')
    skus = inv_repo.list_skus() or []
    sku_map = {s['id']: f"{s['sku_code']} — {s['name']}" for s in skus}
    return render_template('operator/returns.html', shipments=items,
                           clients=clients, client_map=client_map,
                           warehouses=warehouses, locations=locations,
                           skus=skus, sku_map=sku_map, filter_status=status)


@operator_bp.route('/returns/create', methods=['POST'])
@login_required
@_require_operator
def return_create():
    """반품 등록 (고객 회송)."""
    from db_utils import get_repo
    repo = get_repo('order')
    inv_repo = get_repo('inventory')
    client_id = request.form.get('client_id', type=int)
    sku_id = request.form.get('sku_id', type=int)
    quantity = request.form.get('quantity', type=int)
    reason = request.form.get('reason', '').strip()
    location_id = request.form.get('location_id', type=int)

    if not all([client_id, sku_id, quantity]):
        flash('필수 항목을 입력해주세요.', 'warning')
        return redirect(url_for('operator.returns'))

    repo.create_shipment({
        'shipment_type': 'return',
        'client_id': client_id,
        'sku_id': sku_id,
        'quantity': quantity,
        'reason': reason,
        'location_id': location_id,
        'status': 'pending',
    })

    # 재고 복원: movement 기록 + stock 수량 증가
    inv_repo.log_movement({
        'sku_id': sku_id,
        'location_id': location_id,
        'movement_type': 'return_in',
        'quantity': quantity,
        'memo': f'반품입고: {reason}' if reason else '반품입고',
        'user_id': current_user.id,
    })
    if location_id:
        try:
            from services.warehouse_service import _call_rpc
            _call_rpc(inv_repo, 'fn_adjust_stock', {
                'p_operator_id': inv_repo.operator_id,
                'p_sku_id': sku_id,
                'p_location_id': location_id,
                'p_delta': quantity,
                'p_memo': f'반품입고: {reason}' if reason else '반품입고',
                'p_user_id': current_user.id,
            })
        except Exception:
            logger.exception('반품 재고 수량 반영 실패: sku_id=%s, location_id=%s', sku_id, location_id)

    # 반품비 과금
    try:
        from services.client_billing_service import record_return_fee
        record_return_fee(get_repo('client_billing'),
                          get_repo('client_rate'), client_id,
                          quantity=quantity, memo=reason)
    except Exception:
        logger.exception('과금 서비스 호출 자체 실패 (반품): client_id=%s', client_id)
    flash(f'반품 등록 완료 ({quantity}개)', 'success')
    return redirect(url_for('operator.returns'))


# ═══ 창고이동 ═══

@operator_bp.route('/transfers')
@login_required
@_require_operator
def transfers():
    """창고이동 목록."""
    from db_utils import get_repo
    repo = get_repo('order')
    status = request.args.get('status')
    items = repo.list_shipments(status=status, shipment_type='transfer')
    wh_repo = get_repo('warehouse')
    warehouses = wh_repo.list_warehouses() or []
    locations = wh_repo.list_all_locations_with_path()
    inv_repo = get_repo('inventory')
    skus = inv_repo.list_skus() or []
    sku_map = {s['id']: f"{s['sku_code']} — {s['name']}" for s in skus}
    wh_map = {w['id']: w['name'] for w in warehouses}
    return render_template('operator/transfers.html', shipments=items,
                           warehouses=warehouses, locations=locations,
                           skus=skus, sku_map=sku_map, wh_map=wh_map,
                           filter_status=status)


@operator_bp.route('/transfers/create', methods=['POST'])
@login_required
@_require_operator
def transfer_create():
    """창고이동 생성."""
    from db_utils import get_repo
    repo = get_repo('order')
    inv_repo = get_repo('inventory')
    sku_id = request.form.get('sku_id', type=int)
    quantity = request.form.get('quantity', type=int)
    from_loc = request.form.get('from_location_id', type=int)
    to_loc = request.form.get('to_location_id', type=int)
    reason = request.form.get('reason', '').strip()

    if not all([sku_id, quantity]):
        flash('필수 항목을 입력해주세요.', 'warning')
        return redirect(url_for('operator.transfers'))

    if from_loc and to_loc and from_loc == to_loc:
        flash('출발 로케이션과 도착 로케이션이 같습니다.', 'warning')
        return redirect(url_for('operator.transfers'))

    repo.create_shipment({
        'shipment_type': 'transfer',
        'sku_id': sku_id,
        'quantity': quantity,
        'from_warehouse_id': from_loc,
        'to_warehouse_id': to_loc,
        'reason': reason or '창고이동',
        'status': 'completed',
    })

    # RPC 원자적 창고 이동 (동시접속 안전)
    if from_loc and to_loc:
        try:
            from services.warehouse_service import process_transfer
            process_transfer(inv_repo, sku_id, from_loc, to_loc, quantity,
                             memo=reason or '창고이동', user_id=current_user.id)
        except Exception:
            logger.exception('이동 재고 반영 실패')
    elif from_loc:
        try:
            from services.warehouse_service import _call_rpc
            _call_rpc(inv_repo, 'fn_adjust_stock', {
                'p_operator_id': inv_repo.operator_id,
                'p_sku_id': sku_id,
                'p_location_id': from_loc,
                'p_delta': -quantity,
                'p_memo': f'이동출고: {reason}' if reason else '이동출고',
                'p_user_id': current_user.id,
            })
        except Exception:
            logger.exception('이동 출고 재고 반영 실패')
    elif to_loc:
        try:
            from services.warehouse_service import _call_rpc
            _call_rpc(inv_repo, 'fn_adjust_stock', {
                'p_operator_id': inv_repo.operator_id,
                'p_sku_id': sku_id,
                'p_location_id': to_loc,
                'p_delta': quantity,
                'p_memo': f'이동입고: {reason}' if reason else '이동입고',
                'p_user_id': current_user.id,
            })
        except Exception:
            logger.exception('이동 입고 재고 반영 실패')

    flash(f'창고이동 완료 ({quantity}개)', 'success')
    return redirect(url_for('operator.transfers'))


# ═══ 패킹센터 (운영사 뷰) ═══

@operator_bp.route('/packing')
@login_required
@_require_operator
def packing():
    from db_utils import get_repo
    repo = get_repo('packing')
    jobs = repo.list_jobs()
    return render_template('operator/packing.html', jobs=jobs)
