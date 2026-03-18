"""고객사/요금/마켓플레이스/과금정산/고객사 SKU 관련 라우트."""
import logging
from flask import render_template, request, redirect, url_for, flash, abort
from flask_login import login_required, current_user

from . import operator_bp, _require_operator

logger = logging.getLogger(__name__)


def _verify_client_owner(client_id):
    """client_id가 현재 operator 소속인지 검증. 아니면 404."""
    from db_utils import get_repo
    client = get_repo('client').get_client(client_id)
    if not client:
        abort(404)
    return client


# ═══ 고객사관리 ═══

@operator_bp.route('/clients')
@login_required
@_require_operator
def clients():
    from db_utils import get_repo
    repo = get_repo('client')
    items = repo.list_clients()
    return render_template('operator/clients.html', clients=items)


@operator_bp.route('/clients/new', methods=['POST'])
@login_required
@_require_operator
def client_create():
    from db_utils import get_repo
    repo = get_repo('client')
    data = {
        'name': request.form.get('name', '').strip(),
        'business_no': request.form.get('business_no', '').strip(),
        'contact_name': request.form.get('contact_name', '').strip(),
        'contact_phone': request.form.get('contact_phone', '').strip(),
        'contact_email': request.form.get('contact_email', '').strip(),
        'address': request.form.get('address', '').strip(),
        'memo': request.form.get('memo', '').strip(),
    }
    repo.create_client(data)
    flash(f'고객사 "{data["name"]}" 등록 완료', 'success')
    return redirect(url_for('operator.clients'))


@operator_bp.route('/clients/<int:client_id>')
@login_required
@_require_operator
def client_detail(client_id):
    from db_utils import get_repo
    client_repo = get_repo('client')
    client = client_repo.get_client(client_id)
    if not client:
        flash('고객사를 찾을 수 없습니다.', 'warning')
        return redirect(url_for('operator.clients'))
    rate_repo = get_repo('client_rate')
    rates = rate_repo.list_rates(client_id) or []
    rate_count = len(rates)
    mkt_repo = get_repo('client_marketplace')
    mkt_creds = mkt_repo.list_credentials(client_id) or []
    inv_repo = get_repo('inventory')
    client_skus = inv_repo.list_skus(client_id=client_id) or []
    return render_template('operator/client_detail.html', client=client,
                           rates=rates, rate_count=rate_count,
                           mkt_creds=mkt_creds, client_skus=client_skus)


@operator_bp.route('/clients/<int:client_id>/update', methods=['POST'])
@login_required
@_require_operator
def client_update(client_id):
    _verify_client_owner(client_id)
    from db_utils import get_repo
    repo = get_repo('client')
    data = {
        'name': request.form.get('name', '').strip(),
        'business_no': request.form.get('business_no', '').strip(),
        'contact_name': request.form.get('contact_name', '').strip(),
        'contact_phone': request.form.get('contact_phone', '').strip(),
        'contact_email': request.form.get('contact_email', '').strip(),
        'address': request.form.get('address', '').strip(),
        'memo': request.form.get('memo', '').strip(),
    }
    repo.update_client(client_id, data)
    flash('고객사 정보가 수정되었습니다.', 'success')
    return redirect(url_for('operator.client_detail', client_id=client_id))


@operator_bp.route('/clients/<int:client_id>/delete', methods=['POST'])
@login_required
@_require_operator
def client_delete(client_id):
    """고객사 삭제 (soft delete + cascade)."""
    _verify_client_owner(client_id)
    from db_utils import get_repo
    repo = get_repo('client')
    client = repo.get_client(client_id)
    if not client:
        flash('고객사를 찾을 수 없습니다.', 'warning')
        return redirect(url_for('operator.clients'))

    # admin만 삭제 가능
    if not current_user.is_admin():
        flash('고객사 삭제는 관리자만 가능합니다.', 'danger')
        return redirect(url_for('operator.client_detail', client_id=client_id))

    name = client.get('name', '')
    repo.soft_delete_client_cascade(client_id)
    flash(f'고객사 "{name}" 및 연관 데이터가 삭제되었습니다.', 'success')
    return redirect(url_for('operator.clients'))


# ═══ 고객사 요금표 ═══

@operator_bp.route('/clients/<int:client_id>/rates', methods=['POST'])
@login_required
@_require_operator
def client_rate_create(client_id):
    _verify_client_owner(client_id)
    from db_utils import get_repo
    rate_repo = get_repo('client_rate')
    count = rate_repo.count_rates(client_id)
    if count >= 50:
        flash('요금 항목은 최대 50개까지 등록할 수 있습니다.', 'warning')
        return redirect(url_for('operator.client_detail', client_id=client_id))

    data = {
        'client_id': client_id,
        'fee_name': request.form.get('fee_name', '').strip(),
        'fee_type': request.form.get('fee_type', 'fixed'),
        'amount': request.form.get('amount', 0, type=float),
        'unit_label': request.form.get('unit_label', '건').strip(),
        'category': request.form.get('category', 'custom').strip(),
        'memo': request.form.get('memo', '').strip(),
        'sort_order': count,
    }
    if not data['fee_name']:
        flash('항목명을 입력해주세요.', 'warning')
        return redirect(url_for('operator.client_detail', client_id=client_id))

    rate_repo.create_rate(data)
    flash(f'요금 항목 "{data["fee_name"]}" 추가 완료', 'success')
    return redirect(url_for('operator.client_detail', client_id=client_id))


@operator_bp.route('/clients/<int:client_id>/rates/<int:rate_id>/update', methods=['POST'])
@login_required
@_require_operator
def client_rate_update(client_id, rate_id):
    _verify_client_owner(client_id)
    from db_utils import get_repo
    rate_repo = get_repo('client_rate')
    data = {
        'fee_name': request.form.get('fee_name', '').strip(),
        'fee_type': request.form.get('fee_type', 'fixed'),
        'amount': request.form.get('amount', 0, type=float),
        'unit_label': request.form.get('unit_label', '건').strip(),
        'category': request.form.get('category', 'custom').strip(),
        'memo': request.form.get('memo', '').strip(),
    }
    rate_repo.update_rate(rate_id, data)
    flash('요금 항목이 수정되었습니다.', 'success')
    return redirect(url_for('operator.client_detail', client_id=client_id))


@operator_bp.route('/clients/<int:client_id>/rates/<int:rate_id>/delete', methods=['POST'])
@login_required
@_require_operator
def client_rate_delete(client_id, rate_id):
    _verify_client_owner(client_id)
    from db_utils import get_repo
    rate_repo = get_repo('client_rate')
    rate_repo.delete_rate(rate_id)
    flash('요금 항목이 삭제되었습니다.', 'success')
    return redirect(url_for('operator.client_detail', client_id=client_id))


# ═══ 고객사 과금 프리셋 ═══

@operator_bp.route('/clients/<int:client_id>/rates/preset', methods=['POST'])
@login_required
@_require_operator
def client_rate_preset(client_id):
    """프리셋 항목 일괄 추가."""
    _verify_client_owner(client_id)
    from db_utils import get_repo
    from services.client_billing_service import RATE_PRESETS
    rate_repo = get_repo('client_rate')
    categories = request.form.getlist('categories')
    if not categories:
        flash('카테고리를 선택해주세요.', 'warning')
        return redirect(url_for('operator.client_detail', client_id=client_id))

    count = rate_repo.count_rates(client_id)
    existing = rate_repo.list_rates(client_id) or []
    existing_names = {r.get('fee_name') for r in existing}
    added = 0

    for preset in RATE_PRESETS:
        if preset['category'] not in categories:
            continue
        if preset['fee_name'] in existing_names:
            continue
        if count + added >= 50:
            break
        rate_repo.create_rate({
            'client_id': client_id,
            'fee_name': preset['fee_name'],
            'fee_type': preset['fee_type'],
            'amount': preset['amount'],
            'unit_label': preset['unit_label'],
            'category': preset['category'],
            'sort_order': count + added,
        })
        added += 1

    flash(f'프리셋 {added}개 항목 추가 완료', 'success')
    return redirect(url_for('operator.client_detail', client_id=client_id))


# ═══ 고객사 과금 내역/정산 ═══

@operator_bp.route('/clients/<int:client_id>/billing')
@login_required
@_require_operator
def client_billing(client_id):
    """고객사 월별 과금 내역."""
    from db_utils import get_repo
    from services.client_billing_service import CATEGORY_LABELS
    client_repo = get_repo('client')
    client = client_repo.get_client(client_id)
    if not client:
        flash('고객사를 찾을 수 없습니다.', 'warning')
        return redirect(url_for('operator.clients'))

    billing_repo = get_repo('client_billing')
    year_month = request.args.get('month')
    if not year_month:
        from datetime import datetime, timezone
        year_month = datetime.now(timezone.utc).strftime('%Y-%m')

    summary = billing_repo.get_monthly_summary(client_id, year_month)
    invoice = billing_repo.get_invoice(client_id, year_month)

    return render_template('operator/client_billing.html',
                           client=client, year_month=year_month,
                           summary=summary, invoice=invoice,
                           category_labels=CATEGORY_LABELS)


@operator_bp.route('/clients/<int:client_id>/billing/confirm', methods=['POST'])
@login_required
@_require_operator
def client_billing_confirm(client_id):
    """정산서 확정."""
    _verify_client_owner(client_id)
    from db_utils import get_repo
    from datetime import datetime, timezone
    billing_repo = get_repo('client_billing')
    year_month = request.form.get('year_month')
    summary = billing_repo.get_monthly_summary(client_id, year_month)

    invoice = billing_repo.get_invoice(client_id, year_month)
    if invoice:
        billing_repo.update_invoice(invoice['id'], {
            'total_amount': summary['total'],
            'status': 'confirmed',
            'confirmed_at': datetime.now(timezone.utc).isoformat(),
        })
    else:
        billing_repo.create_invoice({
            'client_id': client_id,
            'year_month': year_month,
            'total_amount': summary['total'],
            'status': 'confirmed',
            'confirmed_at': datetime.now(timezone.utc).isoformat(),
        })

    flash(f'{year_month} 정산서 확정 완료 ({summary["total"]:,.0f}원)', 'success')
    return redirect(url_for('operator.client_billing', client_id=client_id,
                            month=year_month))


# ═══ 정산서 Excel 다운로드 ═══

@operator_bp.route('/clients/<int:client_id>/billing/export')
@login_required
@_require_operator
def client_billing_export(client_id):
    """고객사 월별 정산 내역 Excel 다운로드."""
    import io
    from flask import send_file
    from db_utils import get_repo
    from services.client_billing_service import CATEGORY_LABELS
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        flash('openpyxl 미설치', 'danger')
        return redirect(url_for('operator.client_billing', client_id=client_id))

    client_repo = get_repo('client')
    client = client_repo.get_client(client_id)
    if not client:
        flash('고객사를 찾을 수 없습니다.', 'warning')
        return redirect(url_for('operator.clients'))

    billing_repo = get_repo('client_billing')
    year_month = request.args.get('month')
    if not year_month:
        from datetime import datetime, timezone
        year_month = datetime.now(timezone.utc).strftime('%Y-%m')

    summary = billing_repo.get_monthly_summary(client_id, year_month)

    wb = Workbook()
    ws = wb.active
    ws.title = f'{year_month} 정산'

    # 헤더 스타일
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')

    # 타이틀
    ws.merge_cells('A1:G1')
    ws['A1'] = f'{client.get("name", "")} — {year_month} 정산서'
    ws['A1'].font = Font(bold=True, size=14)

    # 카테고리별 소계
    ws.append([])
    ws.append(['카테고리', '합계'])
    for cat_key, cat_amount in summary.get('by_category', {}).items():
        ws.append([CATEGORY_LABELS.get(cat_key, cat_key), f'{cat_amount:,.0f}원'])
    ws.append(['총 합계', f'{summary.get("total", 0):,.0f}원'])
    ws.append([])

    # 상세 내역
    headers = ['카테고리', '항목', '수량', '단가', '금액', '메모', '일시']
    ws.append(headers)
    row_num = ws.max_row
    for col_num, h in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.font = header_font
        cell.fill = header_fill

    for item in summary.get('items', []):
        ws.append([
            CATEGORY_LABELS.get(item.get('category', ''), item.get('category', '')),
            item.get('fee_name', ''),
            item.get('quantity', 0),
            float(item.get('unit_price', 0)),
            float(item.get('total_amount', 0)),
            item.get('memo', ''),
            (item.get('created_at', '')[:10] if item.get('created_at') else ''),
        ])

    # 열 너비 조정
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'{client.get("name", "정산")}_{year_month}_정산서.xlsx'
    return send_file(buf, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ═══ 정산서 PDF 다운로드 ═══

@operator_bp.route('/clients/<int:client_id>/billing/pdf')
@login_required
@_require_operator
def client_billing_pdf(client_id):
    """고객사 월별 정산서 PDF 다운로드."""
    from flask import send_file
    from db_utils import get_repo
    from services.client_billing_service import CATEGORY_LABELS
    try:
        from services.pdf_service import generate_invoice_pdf
    except ImportError:
        flash('PDF 생성 라이브러리(reportlab)가 설치되지 않았습니다.', 'danger')
        return redirect(url_for('operator.client_billing', client_id=client_id))

    client_repo = get_repo('client')
    client = client_repo.get_client(client_id)
    if not client:
        flash('고객사를 찾을 수 없습니다.', 'warning')
        return redirect(url_for('operator.clients'))

    billing_repo = get_repo('client_billing')
    year_month = request.args.get('month')
    if not year_month:
        from datetime import datetime, timezone
        year_month = datetime.now(timezone.utc).strftime('%Y-%m')

    summary = billing_repo.get_monthly_summary(client_id, year_month)
    invoice = billing_repo.get_invoice(client_id, year_month)

    pdf_buf = generate_invoice_pdf(
        client=client,
        year_month=year_month,
        summary=summary,
        invoice=invoice,
    )
    filename = f'{client.get("name", "정산")}_{year_month}_정산서.pdf'
    return send_file(pdf_buf, download_name=filename, mimetype='application/pdf')


# ═══ 보관비 수동 계산 ═══

@operator_bp.route('/clients/<int:client_id>/billing/storage', methods=['POST'])
@login_required
@_require_operator
def client_billing_storage_calc(client_id):
    """보관비 수동 일괄 계산."""
    _verify_client_owner(client_id)
    from db_utils import get_repo
    from services.client_billing_service import calculate_storage_fee
    year_month = request.form.get('year_month')
    force = request.form.get('force') == '1'

    if not year_month:
        from datetime import datetime, timezone
        year_month = datetime.now(timezone.utc).strftime('%Y-%m')

    result = calculate_storage_fee(
        get_repo('client_billing'), get_repo('client_rate'),
        get_repo('inventory'), client_id, year_month, force=force)

    status = result.get('status', 'error') if result else 'error'
    if status == 'ok':
        temp_info = result.get('temp_qty', {})
        detail = ', '.join(f'{k}:{v}개' for k, v in temp_info.items())
        flash(f'{year_month} 보관비 계산 완료 ({detail}, {result.get("days")}일)', 'success')
    elif status == 'already_calculated':
        flash(f'{year_month} 보관비가 이미 계산되어 있습니다 ({result.get("count")}건). '
              f'재계산하려면 [강제 재계산]을 사용하세요.', 'warning')
    elif status == 'no_rates':
        flash('보관비 요금표가 등록되지 않았습니다. 고객사 요금표에서 먼저 등록해주세요.', 'warning')
    else:
        flash(f'보관비 계산 오류: {result.get("error", "알 수 없는 오류")}', 'danger')

    return redirect(url_for('operator.client_billing', client_id=client_id,
                            month=year_month))


# ═══ VAS(부가서비스) 수동 등록 ═══

@operator_bp.route('/clients/<int:client_id>/billing/vas', methods=['POST'])
@login_required
@_require_operator
def client_billing_vas(client_id):
    """VAS 수동 과금 등록."""
    _verify_client_owner(client_id)
    from db_utils import get_repo
    from services.client_billing_service import record_vas_fee
    vas_name = request.form.get('vas_name', '').strip()
    quantity = request.form.get('quantity', 1, type=int)
    memo = request.form.get('memo', '').strip()

    if not vas_name:
        flash('서비스 항목을 선택해주세요.', 'warning')
        return redirect(url_for('operator.client_billing', client_id=client_id))

    record_vas_fee(get_repo('client_billing'), get_repo('client_rate'),
                   client_id, vas_name, quantity=quantity, memo=memo)
    flash(f'VAS 과금 등록: {vas_name} × {quantity}', 'success')
    return redirect(url_for('operator.client_billing', client_id=client_id))


# ═══ 과금 실패 이벤트 재처리 ═══

@operator_bp.route('/billing/failed-events')
@login_required
@_require_operator
def billing_failed_events():
    """과금 실패 이벤트 목록."""
    from db_utils import get_repo
    billing_repo = get_repo('client_billing')
    try:
        events = billing_repo.list_failed_events() or []
    except Exception:
        events = []
    return render_template('operator/billing_failed_events.html', events=events)


@operator_bp.route('/billing/failed-events/<int:event_id>/retry', methods=['POST'])
@login_required
@_require_operator
def billing_failed_event_retry(event_id):
    """실패 이벤트 재처리 시도."""
    import json
    from db_utils import get_repo
    billing_repo = get_repo('client_billing')
    rate_repo = get_repo('client_rate')
    inv_repo = get_repo('inventory')

    event = billing_repo.get_failed_event(event_id)
    if not event:
        flash('이벤트를 찾을 수 없습니다.', 'warning')
        return redirect(url_for('operator.billing_failed_events'))

    event_type = event.get('event_type', '')
    client_id = event.get('client_id')
    event_data = event.get('event_data', '{}')
    if isinstance(event_data, str):
        try:
            event_data = json.loads(event_data)
        except Exception:
            event_data = {}

    try:
        from services import client_billing_service as cbs
        if event_type == 'inbound':
            cbs.record_inbound_fee(billing_repo, rate_repo, client_id, **event_data)
        elif event_type == 'outbound':
            cbs.record_outbound_fee(billing_repo, rate_repo, client_id, **event_data)
        elif event_type == 'material':
            cbs.record_packing_fee(billing_repo, rate_repo, client_id, **event_data)
        elif event_type == 'return':
            cbs.record_return_fee(billing_repo, rate_repo, client_id, **event_data)
        elif event_type == 'storage':
            cbs.calculate_storage_fee(billing_repo, rate_repo, inv_repo,
                                      client_id, **event_data)
        elif event_type == 'vas':
            cbs.record_vas_fee(billing_repo, rate_repo, client_id, **event_data)
        else:
            flash(f'알 수 없는 이벤트 유형: {event_type}', 'warning')
            return redirect(url_for('operator.billing_failed_events'))

        billing_repo.update_failed_event(event_id, {'status': 'resolved'})
        flash(f'재처리 완료: {event_type} (client_id={client_id})', 'success')
    except Exception as e:
        flash(f'재처리 실패: {e}', 'danger')

    return redirect(url_for('operator.billing_failed_events'))


@operator_bp.route('/billing/failed-events/<int:event_id>/dismiss', methods=['POST'])
@login_required
@_require_operator
def billing_failed_event_dismiss(event_id):
    """실패 이벤트 무시/확인 처리."""
    from db_utils import get_repo
    billing_repo = get_repo('client_billing')
    billing_repo.update_failed_event(event_id, {'status': 'dismissed'})
    flash('이벤트가 무시 처리되었습니다.', 'info')
    return redirect(url_for('operator.billing_failed_events'))


# ═══ 고객사 상품 (SKU) ═══

@operator_bp.route('/clients/<int:client_id>/skus', methods=['POST'])
@login_required
@_require_operator
def client_sku_create(client_id):
    """고객사 상품 등록."""
    _verify_client_owner(client_id)
    from db_utils import get_repo
    repo = get_repo('inventory')
    barcode = request.form.get('barcode', '').strip()
    if not barcode:
        flash('바코드는 필수 항목입니다.', 'warning')
        return redirect(url_for('operator.client_detail', client_id=client_id))
    data = {
        'sku_code': request.form.get('sku_code', '').strip(),
        'barcode': barcode,
        'name': request.form.get('name', '').strip(),
        'client_id': client_id,
        'category': request.form.get('category', '').strip(),
        'unit': request.form.get('unit', 'EA').strip(),
        'storage_temp': request.form.get('storage_temp', 'ambient'),
        'weight_g': request.form.get('weight_g', type=float),
        'memo': request.form.get('memo', '').strip(),
    }
    repo.create_sku(data)
    flash(f'상품 "{data["name"]}" 등록 완료', 'success')
    return redirect(url_for('operator.client_detail', client_id=client_id))


@operator_bp.route('/clients/<int:client_id>/skus/<int:sku_id>/update', methods=['POST'])
@login_required
@_require_operator
def client_sku_update(client_id, sku_id):
    """고객사 상품 수정."""
    _verify_client_owner(client_id)
    from db_utils import get_repo
    repo = get_repo('inventory')
    data = {
        'sku_code': request.form.get('sku_code', '').strip(),
        'barcode': request.form.get('barcode', '').strip(),
        'name': request.form.get('name', '').strip(),
        'category': request.form.get('category', '').strip(),
        'unit': request.form.get('unit', 'EA').strip(),
        'storage_temp': request.form.get('storage_temp', 'ambient'),
        'weight_g': request.form.get('weight_g', type=float),
        'memo': request.form.get('memo', '').strip(),
    }
    repo.update_sku(sku_id, data)
    flash('상품 정보가 수정되었습니다.', 'success')
    return redirect(url_for('operator.client_detail', client_id=client_id))


# ═══ 고객사 마켓플레이스 API ═══

@operator_bp.route('/clients/<int:client_id>/marketplace', methods=['POST'])
@login_required
@_require_operator
def client_marketplace_create(client_id):
    """마켓플레이스 API 인증정보 등록."""
    _verify_client_owner(client_id)
    from db_utils import get_repo
    repo = get_repo('client_marketplace')
    channel = request.form.get('channel', '').strip()
    if not channel:
        flash('채널을 선택해주세요.', 'warning')
        return redirect(url_for('operator.client_detail', client_id=client_id))

    data = {
        'client_id': client_id,
        'channel': channel,
        'api_client_id': request.form.get('api_client_id', '').strip(),
        'api_client_secret': request.form.get('api_client_secret', '').strip(),
        'is_active': True,
    }
    extra = {}
    vendor_id = request.form.get('vendor_id', '').strip()
    mall_id = request.form.get('mall_id', '').strip()
    if vendor_id:
        extra['vendor_id'] = vendor_id
    if mall_id:
        extra['mall_id'] = mall_id
    if extra:
        import json
        data['extra_config'] = json.dumps(extra)

    repo.create_credential(data)
    channel_names = {'naver': '네이버', 'coupang': '쿠팡', 'cafe24': '카페24'}
    flash(f'{channel_names.get(channel, channel)} API 등록 완료', 'success')
    return redirect(url_for('operator.client_detail', client_id=client_id))


@operator_bp.route('/clients/<int:client_id>/marketplace/<int:cred_id>/update', methods=['POST'])
@login_required
@_require_operator
def client_marketplace_update(client_id, cred_id):
    """마켓플레이스 API 인증정보 수정."""
    _verify_client_owner(client_id)
    from db_utils import get_repo
    repo = get_repo('client_marketplace')
    data = {
        'api_client_id': request.form.get('api_client_id', '').strip(),
        'api_client_secret': request.form.get('api_client_secret', '').strip(),
        'is_active': request.form.get('is_active') == 'on',
    }
    extra = {}
    vendor_id = request.form.get('vendor_id', '').strip()
    mall_id = request.form.get('mall_id', '').strip()
    if vendor_id:
        extra['vendor_id'] = vendor_id
    if mall_id:
        extra['mall_id'] = mall_id
    if extra:
        import json
        data['extra_config'] = json.dumps(extra)

    repo.update_credential(cred_id, data)
    flash('API 정보가 수정되었습니다.', 'success')
    return redirect(url_for('operator.client_detail', client_id=client_id))


@operator_bp.route('/clients/<int:client_id>/marketplace/<int:cred_id>/delete', methods=['POST'])
@login_required
@_require_operator
def client_marketplace_delete(client_id, cred_id):
    """마켓플레이스 API 인증정보 삭제."""
    _verify_client_owner(client_id)
    from db_utils import get_repo
    repo = get_repo('client_marketplace')
    repo.delete_credential(cred_id)
    flash('API 연동이 삭제되었습니다.', 'success')
    return redirect(url_for('operator.client_detail', client_id=client_id))
