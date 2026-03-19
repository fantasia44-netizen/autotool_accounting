"""재고/입고/조정/수불부/SKU 관련 라우트."""
import logging
import uuid
from datetime import datetime
from flask import render_template, request, redirect, url_for, flash, jsonify
from flask_login import login_required, current_user

from . import operator_bp, _require_operator

logger = logging.getLogger(__name__)


# ═══ 재고현황 ═══

@operator_bp.route('/inventory')
@login_required
@_require_operator
def inventory():
    from db_utils import get_repo
    repo = get_repo('inventory')
    client_id = request.args.get('client_id', type=int)
    search = request.args.get('search')
    stocks = repo.list_all_stock()
    skus = repo.list_skus(client_id=client_id, search=search)
    client_repo = get_repo('client')
    clients = client_repo.list_clients()
    return render_template('operator/inventory.html', stocks=stocks, skus=skus,
                           clients=clients, filter_client_id=client_id,
                           filter_search=search)


# ═══ 입고관리 ═══

@operator_bp.route('/inbound', methods=['GET', 'POST'])
@login_required
@_require_operator
def inbound():
    from db_utils import get_repo
    inv_repo = get_repo('inventory')
    wh_repo = get_repo('warehouse')

    if request.method == 'POST':
        from services.warehouse_service import process_inbound
        sku_id = request.form.get('sku_id', type=int)
        location_id = request.form.get('location_id', type=int)
        quantity = request.form.get('quantity', type=int)
        lot_number = request.form.get('lot_number', '').strip() or None
        expiry_date = request.form.get('expiry_date', '').strip() or None
        storage_temp = request.form.get('storage_temp', '').strip() or None
        memo = request.form.get('memo', '').strip()

        try:
            process_inbound(inv_repo, sku_id, location_id, quantity,
                            lot_number=lot_number, memo=memo,
                            user_id=current_user.id,
                            expiry_date=expiry_date)
            flash(f'{quantity}개 입고 완료', 'success')
            # 과금 기록 (서비스 내부에서 DLQ 처리)
            try:
                sku = inv_repo.get_sku(sku_id)
                cid = sku.get('client_id') if sku else None
                if cid:
                    from services.client_billing_service import record_inbound_fee
                    record_inbound_fee(get_repo('client_billing'),
                                       get_repo('client_rate'), cid,
                                       quantity=quantity, memo=memo)
            except Exception:
                logger.exception('과금 서비스 호출 자체 실패 (입고): sku_id=%s', sku_id)
        except Exception as e:
            flash(f'입고 오류: {e}', 'danger')
        return redirect(url_for('operator.inbound'))

    client_repo = get_repo('client')
    clients = client_repo.list_clients() or []
    skus = inv_repo.list_skus()
    locations = wh_repo.list_all_locations_with_path()
    warehouses = wh_repo.list_warehouses()
    recent = inv_repo.list_movements(movement_type='inbound', limit=20)
    sku_map = {s['id']: f"{s['sku_code']} — {s['name']}" for s in skus}
    # 로케이션 ID→표시명 맵 (최근입고에서 사용)
    loc_map = {loc['id']: loc.get('display', loc.get('code', '?')) for loc in locations}
    return render_template('operator/inbound.html', skus=skus, locations=locations,
                           warehouses=warehouses, recent_inbounds=recent,
                           clients=clients, sku_map=sku_map, loc_map=loc_map)


# ═══ API ═══

@operator_bp.route('/api/skus-by-client')
@login_required
@_require_operator
def api_skus_by_client():
    """고객사별 SKU 목록 JSON API."""
    from flask import jsonify
    from db_utils import get_repo
    client_id = request.args.get('client_id', type=int)
    if not client_id:
        return jsonify([])
    # 테넌트 격리: 요청된 client_id가 현재 운영자 소속인지 검증
    client_repo = get_repo('client')
    client = client_repo.get_client(client_id)
    if not client:
        return jsonify([])
    repo = get_repo('inventory')
    skus = repo.list_skus(client_id=client_id) or []
    return jsonify([{'id': s['id'], 'sku_code': s['sku_code'], 'name': s['name'],
                     'barcode': s.get('barcode', '')} for s in skus])


# ═══ 재고조정 ═══

@operator_bp.route('/adjustment', methods=['GET', 'POST'])
@login_required
@_require_operator
def adjustment():
    from db_utils import get_repo
    inv_repo = get_repo('inventory')

    if request.method == 'POST':
        sku_id = request.form.get('sku_id', type=int)
        location_id = request.form.get('location_id', type=int)
        delta = request.form.get('delta', type=int)
        lot_number = request.form.get('lot_number', '').strip() or None
        memo = request.form.get('memo', '').strip()

        # RPC 원자적 재고 조정 (동시접속 안전)
        from services.warehouse_service import _call_rpc
        result = _call_rpc(inv_repo, 'fn_adjust_stock', {
            'p_operator_id': inv_repo.operator_id,
            'p_sku_id': sku_id,
            'p_location_id': location_id,
            'p_delta': delta,
            'p_lot_number': lot_number,
            'p_memo': memo or '재고 조정',
            'p_user_id': current_user.id,
        })
        if not result.get('ok'):
            flash(f"조정 실패: {result.get('error', '알 수 없는 오류')}", 'danger')
            return redirect(url_for('operator.adjustment'))
        flash(f'재고 조정 완료 ({delta:+d})', 'success')
        return redirect(url_for('operator.adjustment'))

    wh_repo = get_repo('warehouse')
    skus = inv_repo.list_skus()
    locations = wh_repo.list_all_locations_with_path()
    warehouses = wh_repo.list_warehouses()
    recent = inv_repo.list_movements(movement_type='adjust', limit=20)
    loc_map = {loc['id']: loc.get('display', loc.get('code', '?')) for loc in locations}
    sku_map = {s['id']: f"{s['sku_code']} — {s['name']}" for s in skus}
    return render_template('operator/adjustment.html', skus=skus, locations=locations,
                           warehouses=warehouses, recent_adjustments=recent,
                           loc_map=loc_map, sku_map=sku_map)


# ═══ 재고실사 엑셀 일괄조정 ═══

@operator_bp.route('/adjustment/sample-excel')
@login_required
@_require_operator
def adjustment_sample_excel():
    """재고실사 샘플 엑셀 다운로드."""
    import io
    from flask import send_file
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        flash('openpyxl 미설치', 'danger')
        return redirect(url_for('operator.adjustment'))

    wb = Workbook()
    ws = wb.active
    ws.title = '재고실사'

    # 헤더 스타일
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=10)

    headers = ['SKU코드(필수)', '바코드', '품명(자동)', '로케이션코드', '실사수량(필수)', 'LOT번호']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # 샘플 데이터
    ws.append(['SKU-001', '8801234567890', '', '', 85, ''])
    ws.append(['SKU-002', '8801234567891', '', '', 120, 'LOT-2024A'])

    # 컬럼 너비
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name='재고실사_양식.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@operator_bp.route('/adjustment/bulk-preview', methods=['POST'])
@login_required
@_require_operator
def adjustment_bulk_preview():
    """엑셀 업로드 → 미리보기 JSON 반환 (AJAX)."""
    from db_utils import get_repo
    try:
        from openpyxl import load_workbook
    except ImportError:
        return jsonify({'ok': False, 'error': 'openpyxl 미설치'}), 500

    f = request.files.get('file')
    if not f or not f.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'ok': False, 'error': '엑셀 파일(.xlsx)을 선택해주세요.'})

    location_id = request.form.get('location_id', type=int)
    survey_date = request.form.get('survey_date', '').strip()  # 실사 기준일 (YYYY-MM-DD)

    inv_repo = get_repo('inventory')
    skus = inv_repo.list_skus() or []
    sku_by_code = {s['sku_code']: s for s in skus}
    sku_by_barcode = {s.get('barcode', ''): s for s in skus if s.get('barcode')}
    stocks = inv_repo.list_all_stock() or []

    # SKU+로케이션별 현재고 맵
    stock_map = {}
    for st in stocks:
        key = (st['sku_id'], st.get('location_id'))
        stock_map[key] = stock_map.get(key, 0) + st.get('quantity', 0)

    # SKU별 전체 현재고 (로케이션 무관)
    stock_total = {}
    for st in stocks:
        sid = st['sku_id']
        stock_total[sid] = stock_total.get(sid, 0) + st.get('quantity', 0)

    # 기준일 지정 시: 기준일 이후 movement를 역산하여 기준일 시점 재고 계산
    # "기준일 시점 재고 = 현재 재고 - (기준일 다음날~오늘 movement 합계)"
    after_survey_delta = {}  # {(sku_id, location_id): 이후 변동 합계}
    after_survey_total = {}  # {sku_id: 이후 변동 합계}
    if survey_date:
        # 기준일 다음날 00:00부터 오늘까지의 모든 movement
        survey_next = survey_date + 'T23:59:59'
        all_movements = inv_repo.list_movements(date_from=survey_next, limit=10000)
        for mv in all_movements:
            sid = mv.get('sku_id')
            lid = mv.get('location_id')
            qty = mv.get('quantity', 0)
            key = (sid, lid)
            after_survey_delta[key] = after_survey_delta.get(key, 0) + qty
            after_survey_total[sid] = after_survey_total.get(sid, 0) + qty

    wb = load_workbook(f, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    preview = []
    errors = []
    for i, row in enumerate(rows, 2):
        if not row or (not row[0] and (len(row) < 2 or not row[1])):
            continue
        sku_code = str(row[0] or '').strip()
        barcode = str(row[1] or '').strip() if len(row) > 1 else ''
        loc_code = str(row[3] or '').strip() if len(row) > 3 else ''
        lot_number = str(row[5] or '').strip() if len(row) > 5 else ''

        # 실사수량
        try:
            actual_qty = int(row[4]) if len(row) > 4 and row[4] is not None else None
        except (ValueError, TypeError):
            errors.append(f'행 {i}: 실사수량이 숫자가 아닙니다.')
            continue
        if actual_qty is None:
            errors.append(f'행 {i}: 실사수량이 비어있습니다.')
            continue

        # SKU 찾기 (코드 → 바코드 순)
        sku = sku_by_code.get(sku_code)
        if not sku and barcode:
            sku = sku_by_barcode.get(barcode)
        if not sku:
            errors.append(f'행 {i}: SKU "{sku_code or barcode}" 을(를) 찾을 수 없습니다.')
            continue

        # 현재고 (로케이션 지정 시 해당 위치, 아니면 전체)
        target_loc = location_id
        if loc_code:
            # 로케이션 코드로 ID 찾기
            wh_repo = get_repo('warehouse')
            locs = wh_repo.list_all_locations_with_path()
            loc_match = next((l for l in locs if l.get('code') == loc_code), None)
            if loc_match:
                target_loc = loc_match['id']

        if target_loc:
            now_qty = stock_map.get((sku['id'], target_loc), 0)
            after_delta = after_survey_delta.get((sku['id'], target_loc), 0)
        else:
            now_qty = stock_total.get(sku['id'], 0)
            after_delta = after_survey_total.get(sku['id'], 0)

        # 기준일 시점 시스템 재고 = 현재 재고 - (기준일 이후 변동)
        system_qty_at_date = now_qty - after_delta if survey_date else now_qty
        # 차이 = 실사수량 - 기준일시점재고 (입출고는 그대로 유지)
        delta = actual_qty - system_qty_at_date

        preview.append({
            'row': i,
            'sku_id': sku['id'],
            'sku_code': sku['sku_code'],
            'sku_name': sku.get('name', ''),
            'location_id': target_loc,
            'lot_number': lot_number or None,
            'current_qty': now_qty,
            'system_qty_at_date': system_qty_at_date,
            'after_movements': after_delta,
            'actual_qty': actual_qty,
            'delta': delta,
            'survey_date': survey_date or None,
        })

    return jsonify({
        'ok': True,
        'preview': preview,
        'errors': errors,
        'total_items': len(preview),
        'increase_count': sum(1 for p in preview if p['delta'] > 0),
        'decrease_count': sum(1 for p in preview if p['delta'] < 0),
        'no_change_count': sum(1 for p in preview if p['delta'] == 0),
    })


@operator_bp.route('/adjustment/bulk-apply', methods=['POST'])
@login_required
@_require_operator
def adjustment_bulk_apply():
    """미리보기 확인 후 일괄 적용."""
    import json
    from db_utils import get_repo
    from services.warehouse_service import _call_rpc

    data = request.get_json()
    if not data or 'items' not in data:
        return jsonify({'ok': False, 'error': '적용할 데이터가 없습니다.'})

    items = data['items']
    memo = data.get('memo', '재고실사 일괄조정')
    batch_id = f"ADJ-{datetime.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:6].upper()}"

    inv_repo = get_repo('inventory')
    results = {'ok': 0, 'fail': 0, 'skipped': 0, 'details': []}

    for item in items:
        delta = item.get('delta', 0)
        if delta == 0:
            results['skipped'] += 1
            continue

        sku_id = item['sku_id']
        location_id = item.get('location_id')
        lot_number = item.get('lot_number')
        batch_memo = f"[{batch_id}] {memo} (실사:{item.get('actual_qty')} 시스템:{item.get('current_qty')} 차이:{delta:+d})"

        try:
            result = _call_rpc(inv_repo, 'fn_adjust_stock', {
                'p_operator_id': inv_repo.operator_id,
                'p_sku_id': sku_id,
                'p_location_id': location_id,
                'p_delta': delta,
                'p_lot_number': lot_number,
                'p_memo': batch_memo,
                'p_user_id': current_user.id,
            })
            if result.get('ok'):
                results['ok'] += 1
                results['details'].append({
                    'sku_code': item.get('sku_code', ''),
                    'delta': delta,
                    'new_qty': result.get('new_quantity'),
                    'status': 'success',
                })
            else:
                results['fail'] += 1
                results['details'].append({
                    'sku_code': item.get('sku_code', ''),
                    'delta': delta,
                    'error': result.get('error', ''),
                    'status': 'fail',
                })
        except Exception as e:
            results['fail'] += 1
            results['details'].append({
                'sku_code': item.get('sku_code', ''),
                'delta': delta,
                'error': str(e),
                'status': 'fail',
            })

    return jsonify({
        'ok': True,
        'batch_id': batch_id,
        'success': results['ok'],
        'fail': results['fail'],
        'skipped': results['skipped'],
        'details': results['details'],
    })


@operator_bp.route('/adjustment/batch-history')
@login_required
@_require_operator
def adjustment_batch_history():
    """배치 조정 이력 조회 (JSON)."""
    from db_utils import get_repo
    inv_repo = get_repo('inventory')
    # ADJ- 로 시작하는 메모의 movement 조회
    movements = inv_repo.list_movements(movement_type='adjust', limit=500)
    # 배치별 그룹핑
    batches = {}
    for m in movements:
        memo = m.get('memo', '')
        if not memo.startswith('[ADJ-'):
            continue
        bid = memo.split(']')[0][1:]  # [ADJ-xxx] → ADJ-xxx
        if bid not in batches:
            batches[bid] = {
                'batch_id': bid,
                'created_at': m.get('created_at', ''),
                'items': [],
                'total_increase': 0,
                'total_decrease': 0,
            }
        batches[bid]['items'].append({
            'movement_id': m.get('id'),
            'sku_id': m.get('sku_id'),
            'location_id': m.get('location_id'),
            'quantity': m.get('quantity', 0),
            'lot_number': m.get('lot_number'),
            'memo': memo,
        })
        qty = m.get('quantity', 0)
        if qty > 0:
            batches[bid]['total_increase'] += qty
        else:
            batches[bid]['total_decrease'] += qty

    batch_list = sorted(batches.values(), key=lambda b: b['created_at'], reverse=True)
    return jsonify({'ok': True, 'batches': batch_list[:20]})


@operator_bp.route('/adjustment/batch-rollback', methods=['POST'])
@login_required
@_require_operator
def adjustment_batch_rollback():
    """배치 되돌리기: 동일 배치의 모든 조정을 역방향 적용."""
    import json
    from db_utils import get_repo
    from services.warehouse_service import _call_rpc

    data = request.get_json()
    batch_id = data.get('batch_id')
    if not batch_id:
        return jsonify({'ok': False, 'error': '배치 ID가 필요합니다.'})

    inv_repo = get_repo('inventory')
    movements = inv_repo.list_movements(movement_type='adjust', limit=1000)

    # 해당 배치 항목 필터
    batch_items = [m for m in movements
                   if m.get('memo', '').startswith(f'[{batch_id}]')]
    if not batch_items:
        return jsonify({'ok': False, 'error': f'배치 "{batch_id}" 이력을 찾을 수 없습니다.'})

    # 이미 되돌린 배치인지 확인
    rollback_check = [m for m in movements
                      if m.get('memo', '').startswith(f'[ROLLBACK-{batch_id}]')]
    if rollback_check:
        return jsonify({'ok': False, 'error': f'이미 되돌린 배치입니다.'})

    rollback_memo_prefix = f"[ROLLBACK-{batch_id}]"
    results = {'ok': 0, 'fail': 0}

    for item in batch_items:
        reverse_delta = -item.get('quantity', 0)
        if reverse_delta == 0:
            continue
        try:
            result = _call_rpc(inv_repo, 'fn_adjust_stock', {
                'p_operator_id': inv_repo.operator_id,
                'p_sku_id': item['sku_id'],
                'p_location_id': item.get('location_id'),
                'p_delta': reverse_delta,
                'p_lot_number': item.get('lot_number'),
                'p_memo': f"{rollback_memo_prefix} 되돌리기 ({reverse_delta:+d})",
                'p_user_id': current_user.id,
            })
            if result.get('ok'):
                results['ok'] += 1
            else:
                results['fail'] += 1
        except Exception:
            results['fail'] += 1

    return jsonify({
        'ok': True,
        'batch_id': batch_id,
        'rollback_success': results['ok'],
        'rollback_fail': results['fail'],
    })


# ═══ 수불부 ═══

@operator_bp.route('/ledger')
@login_required
@_require_operator
def ledger():
    from db_utils import get_repo
    repo = get_repo('inventory')
    sku_id = request.args.get('sku_id', type=int)
    movement_type = request.args.get('type')
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    movements = repo.list_movements(sku_id=sku_id, movement_type=movement_type,
                                    date_from=date_from, date_to=date_to)
    skus = repo.list_skus()
    wh_repo = get_repo('warehouse')
    locations = wh_repo.list_all_locations_with_path()
    loc_map = {loc['id']: loc.get('display', loc.get('code', '?')) for loc in locations}
    sku_map = {s['id']: f"{s['sku_code']} — {s['name']}" for s in skus}
    return render_template('operator/ledger.html', movements=movements, skus=skus,
                           loc_map=loc_map, sku_map=sku_map,
                           filter_sku_id=sku_id, filter_type=movement_type,
                           filter_date_from=date_from, filter_date_to=date_to)


@operator_bp.route('/inventory/export')
@login_required
@_require_operator
def inventory_export():
    """재고 현황 엑셀 다운로드."""
    import io
    from flask import send_file
    from db_utils import get_repo
    try:
        from openpyxl import Workbook
    except ImportError:
        flash('openpyxl 미설치', 'danger')
        return redirect(url_for('operator.inventory'))

    inv_repo = get_repo('inventory')
    skus = inv_repo.list_skus() or []
    stocks = inv_repo.list_all_stock() or []
    sku_stock = {}
    for st in stocks:
        sid = st.get('sku_id')
        sku_stock[sid] = sku_stock.get(sid, 0) + st.get('quantity', 0)

    wb = Workbook()
    ws = wb.active
    ws.title = '재고현황'
    ws.append(['SKU코드', '바코드', '품명', '카테고리', '보관온도', '현재고', '최소재고'])
    for s in skus:
        ws.append([
            s.get('sku_code', ''), s.get('barcode', ''), s.get('name', ''),
            s.get('category', ''), s.get('storage_temp', 'ambient'),
            sku_stock.get(s['id'], 0), s.get('min_stock_qty', 0),
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name='재고현황.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@operator_bp.route('/ledger/export')
@login_required
@_require_operator
def ledger_export():
    """수불장 엑셀 다운로드."""
    import io
    from flask import send_file
    from db_utils import get_repo
    try:
        from openpyxl import Workbook
    except ImportError:
        flash('openpyxl 미설치', 'danger')
        return redirect(url_for('operator.ledger'))

    repo = get_repo('inventory')
    sku_id = request.args.get('sku_id', type=int)
    movement_type = request.args.get('type')
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    movements = repo.list_movements(sku_id=sku_id, movement_type=movement_type,
                                    date_from=date_from, date_to=date_to)

    wb = Workbook()
    ws = wb.active
    ws.title = '수불장'
    ws.append(['일시', '유형', 'SKU ID', '수량', '로케이션', 'LOT', '메모'])
    for m in movements:
        ws.append([
            m.get('created_at', '')[:19] if m.get('created_at') else '',
            m.get('movement_type', ''),
            m.get('sku_id', ''),
            m.get('quantity', 0),
            m.get('location_id', ''),
            m.get('lot_number', ''),
            m.get('memo', ''),
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name='수불장.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ═══ 상품마스터 (SKU) ═══

@operator_bp.route('/skus')
@login_required
@_require_operator
def skus():
    from db_utils import get_repo
    repo = get_repo('inventory')
    search = request.args.get('search')
    client_id = request.args.get('client_id', type=int)
    items = repo.list_skus(client_id=client_id, search=search)
    client_repo = get_repo('client')
    clients = client_repo.list_clients() or []
    client_map = {c['id']: c['name'] for c in clients}
    return render_template('operator/skus.html', skus=items, clients=clients,
                           client_map=client_map,
                           filter_search=search, filter_client_id=client_id)


@operator_bp.route('/skus/sample-excel')
@login_required
@_require_operator
def sku_sample_excel():
    """상품 등록 샘플 엑셀 다운로드."""
    import io
    from flask import send_file
    try:
        from openpyxl import Workbook
    except ImportError:
        flash('openpyxl 패키지가 설치되지 않았습니다.', 'danger')
        return redirect(url_for('operator.skus'))

    wb = Workbook()
    ws = wb.active
    ws.title = '상품등록'
    headers = ['sku_code(필수)', 'barcode(필수)', 'name(필수)', 'category',
               'unit(EA/BOX/PALLET/PACK/SET)', 'storage_temp(ambient/cold/frozen)',
               'weight_g', 'memo']
    ws.append(headers)
    ws.append(['SKU-001', '8801234567890', '테스트상품A', '식품', 'EA', 'ambient', 500, '샘플'])
    ws.append(['SKU-002', '8801234567891', '냉동상품B', '냉동식품', 'BOX', 'frozen', 2000, ''])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name='상품등록_샘플.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@operator_bp.route('/skus/bulk-upload', methods=['POST'])
@login_required
@_require_operator
def sku_bulk_upload():
    """상품 엑셀 일괄 업로드."""
    from db_utils import get_repo
    try:
        from openpyxl import load_workbook
    except ImportError:
        flash('openpyxl 패키지가 설치되지 않았습니다.', 'danger')
        return redirect(url_for('operator.skus'))

    f = request.files.get('file')
    if not f or not f.filename.endswith(('.xlsx', '.xls')):
        flash('엑셀 파일(.xlsx)을 선택해주세요.', 'warning')
        return redirect(url_for('operator.skus'))
    # MIME 타입 검증
    allowed_mimes = (
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'application/vnd.ms-excel',
        'application/octet-stream',  # 일부 브라우저
    )
    if f.content_type not in allowed_mimes:
        flash('허용되지 않는 파일 형식입니다.', 'warning')
        return redirect(url_for('operator.skus'))

    client_id = request.form.get('client_id', type=int)
    repo = get_repo('inventory')

    wb = load_workbook(f, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    ok, fail = 0, 0
    for row in rows:
        if not row or len(row) < 3:
            fail += 1
            continue
        sku_code = str(row[0] or '').strip()
        barcode = str(row[1] or '').strip()
        name = str(row[2] or '').strip()
        if not sku_code or not barcode or not name:
            fail += 1
            continue
        category = str(row[3] or '').strip() if len(row) > 3 else ''
        unit = str(row[4] or 'EA').strip().upper() if len(row) > 4 else 'EA'
        if unit not in ('EA', 'BOX', 'PALLET', 'PACK', 'SET'):
            unit = 'EA'
        storage_temp = str(row[5] or 'ambient').strip().lower() if len(row) > 5 else 'ambient'
        if storage_temp not in ('ambient', 'cold', 'frozen'):
            storage_temp = 'ambient'
        weight_g = None
        if len(row) > 6 and row[6]:
            try:
                weight_g = float(row[6])
            except (ValueError, TypeError):
                pass
        memo = str(row[7] or '').strip() if len(row) > 7 else ''

        data = {
            'sku_code': sku_code, 'barcode': barcode, 'name': name,
            'category': category, 'unit': unit, 'storage_temp': storage_temp,
            'weight_g': weight_g, 'memo': memo,
        }
        if client_id:
            data['client_id'] = client_id
        try:
            repo.create_sku(data)
            ok += 1
        except Exception:
            fail += 1

    flash(f'업로드 완료: 성공 {ok}건, 실패 {fail}건', 'success' if ok else 'warning')
    if client_id:
        return redirect(url_for('operator.client_detail', client_id=client_id))
    return redirect(url_for('operator.skus'))


@operator_bp.route('/skus/new', methods=['POST'])
@login_required
@_require_operator
def sku_create():
    from db_utils import get_repo
    repo = get_repo('inventory')
    barcode = request.form.get('barcode', '').strip()
    if not barcode:
        flash('바코드는 필수 항목입니다.', 'warning')
        return redirect(url_for('operator.skus'))
    data = {
        'sku_code': request.form.get('sku_code', '').strip(),
        'barcode': barcode,
        'name': request.form.get('name', '').strip(),
        'client_id': request.form.get('client_id', type=int),
        'category': request.form.get('category', '').strip(),
        'unit': request.form.get('unit', 'EA').strip(),
        'storage_temp': request.form.get('storage_temp', 'ambient'),
        'weight_g': request.form.get('weight_g', type=float),
        'memo': request.form.get('memo', '').strip(),
    }
    repo.create_sku(data)
    flash(f'상품 "{data["name"]}" 등록 완료', 'success')
    return redirect(url_for('operator.skus'))
